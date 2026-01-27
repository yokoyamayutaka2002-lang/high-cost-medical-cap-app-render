#!/usr/bin/env python3
"""Find rows in topical Excel whose 品名 cell exactly equals a target string.

Prints: sheet, row, repr(品名), repr(規格), 薬価
"""
from pathlib import Path
from openpyxl import load_workbook


def find_header_indices(header_row):
    names = [ (h or '') for h in header_row ]
    idx = {'name': None, 'strength': None, 'price': None}
    name_keys = ['品名', '医薬品名', '薬品名', '商品名', '製品名']
    strength_keys = ['規格', '含量', '規格・含量']
    price_keys = ['薬価', '価格', '単価']
    for i, h in enumerate(names):
        if not h:
            continue
        for k in name_keys:
            if k in h:
                idx['name'] = i
        for k in strength_keys:
            if k in h:
                idx['strength'] = i
        for k in price_keys:
            if k in h:
                idx['price'] = i
    return idx


def main():
    repo = Path('.')
    src = repo / 'data' / 'source_excel'
    # prefer exact filename, otherwise pick file containing 'topical'
    candidates = [p for p in src.iterdir() if p.is_file() and 'topical' in p.name.lower()]
    if not candidates:
        print('ERROR: no topical workbook found in', src)
        return
    xlsx = candidates[0]
    print('Using workbook:', xlsx.name)

    target = 'アドエア５００ディスカス６０吸入用'

    wb = load_workbook(filename=str(xlsx), data_only=True)
    found_any = False
    for sheet in wb.worksheets:
        try:
            header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            continue
        idx = find_header_indices(header)
        if idx['name'] is None:
            continue

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            cell_name = row[idx['name']]
            if cell_name is None:
                continue
            # exact match (no normalization)
            if str(cell_name).strip() == target:
                found_any = True
                spec = ''
                if idx.get('strength') is not None:
                    spec = row[idx['strength']]
                price = ''
                if idx.get('price') is not None:
                    price = row[idx['price']]
                print('FOUND: sheet=%s, row=%s, name=%r, spec=%r, price=%r' % (sheet.title, row_idx, str(cell_name), spec, price))

    if not found_any:
        print('No exact match found for target:', target)


if __name__ == '__main__':
    main()
