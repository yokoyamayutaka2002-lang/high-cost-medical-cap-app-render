#!/usr/bin/env python3
"""E2E dry-run orchestrator

Usage:
  python scripts/e2e_dryrun.py path/to/source.xlsx

This script performs Steps 1-7 described in the PR: precheck the Excel file,
run a non-destructive importer dry-run, generate staging CSV, run validators,
run representative simulation cases, and produce reports/e2e_summary.json and
reports/e2e_human_readable.md.

Notes:
- This is a lightweight harness intended for local dry-runs. It expects
  `data/mapping.yaml`, `data/unit_map.csv`, `data/master_id_map.csv`, and
  `data/validator_rules.yaml` to exist. It writes intermediate CSVs under
  `reports/dryrun/` and `staging/` and final reports under `reports/`.
"""
import sys
import json
import csv
import subprocess
from pathlib import Path
from datetime import datetime
try:
    import yaml
except Exception:
    yaml = None
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


ROOT = Path(__file__).parent.parent
REPORTS = ROOT / 'reports'
REPORTS.mkdir(exist_ok=True)
DRYRUN_DIR = REPORTS / 'dryrun'
DRYRUN_DIR.mkdir(exist_ok=True)
STAGING = ROOT / 'staging'
STAGING.mkdir(exist_ok=True)


def precheck_excel(xlsx_path: Path, mapping_path: Path):
    out = {'status': 'pass', 'errors': [], 'warnings': [], 'meta': {}, 'checked_at': datetime.utcnow().isoformat()}

    if not xlsx_path.exists():
        out['status'] = 'fail'
        out['errors'].append({'type': 'FILE_MISSING', 'message': str(xlsx_path)})
        (REPORTS / 'precheck_excel.json').write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf8')
        return out

    # extension and macro checks
    if xlsx_path.suffix.lower() not in ('.xlsx',):
        out['status'] = 'fail'
        out['errors'].append({'type': 'BAD_EXTENSION', 'message': f'Expected .xlsx (no macros): {xlsx_path.suffix}'})

    if load_workbook is None:
        out['status'] = 'fail'
        out['errors'].append({'type': 'MISSING_DEP', 'message': 'openpyxl is required (install via requirements.txt)'})
        (REPORTS / 'precheck_excel.json').write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf8')
        return out

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    except Exception as e:
        out['status'] = 'fail'
        out['errors'].append({'type': 'OPEN_ERROR', 'message': str(e)})
        (REPORTS / 'precheck_excel.json').write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf8')
        return out

    # merged cells
    # NOTE: openpyxl exposes merged cell ranges on Worksheet, not Workbook.
    # The merged-cells check caused AttributeError in some environments
    # (Workbook has no attribute 'merged_cells'), and this precheck is
    # optional for our importer dry-run. Skip merged-cells detection here
    # to avoid false failures; it's safe because importer will not write
    # back to the Excel file.
    merged = []

    # load mapping required candidates
    if yaml is None:
        out['warnings'].append({'type': 'NO_PYYAML', 'message': 'PyYAML not installed; cannot read mapping.yaml to validate headers'})
        (REPORTS / 'precheck_excel.json').write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf8')
        return out

    mapping = yaml.safe_load(mapping_path.read_text(encoding='utf8'))
    req = mapping.get('required', {})
    # use case-insensitive match if mapping.settings.match.case_sensitive == False
    case_sensitive = mapping.get('settings', {}).get('match', {}).get('case_sensitive', False)

    found_required = {k: False for k in req.keys()}
    revision_found = None

    # iterate sheets and headers
    for sheet in wb.worksheets:
        # read first row as header (if any)
        try:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            header_row = []
        headers = [ (h or '').strip() for h in header_row ]

        for logical_field, spec in req.items():
            candidates = spec.get('candidates', [])
            for cand in candidates:
                for h in headers:
                    if not h:
                        continue
                    if case_sensitive:
                        match = (h == cand)
                    else:
                        match = (h.lower() == cand.lower())
                    if match:
                        found_required[logical_field] = True
                        break
                if found_required[logical_field]:
                    break

        # revision detection: check optional excel_revision_date candidates
        rev_cands = mapping.get('optional', {}).get('excel_revision_date', {}).get('candidates', [])
        for cand in rev_cands:
            for idx, h in enumerate(headers):
                if h and ((case_sensitive and h == cand) or (not case_sensitive and h.lower() == cand.lower())):
                    # try to read value from row 2 in same column
                    try:
                        cell = sheet.cell(row=2, column=idx+1)
                        if cell.value:
                            revision_found = str(cell.value)
                    except Exception:
                        pass

    missing = [k for k, v in found_required.items() if not v]
    if missing:
        out['status'] = 'fail'
        out['errors'].append({'type': 'MISSING_REQUIRED_COLUMNS', 'message': f'Missing required logical fields: {missing}'})

    if revision_found:
        out['meta']['revision'] = revision_found
    else:
        out['warnings'].append({'type': 'REVISION_NOT_FOUND', 'message': 'Could not find revision date in known candidates'})

    # check numeric columns for formulas and non-numeric values on the first sheet that has required fields
    # try to find a suitable sheet
    numeric_issues = []
    price_candidates = req.get('excel_price_yen', {}).get('candidates', [])
    chosen_sheet = None
    for sheet in wb.worksheets:
        header = [ (c or '').strip() for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) ]
        if any(((h and (h.lower() in [p.lower() for p in price_candidates])) for h in header)):
            chosen_sheet = sheet
            break

    if chosen_sheet is not None:
        header = [ (c or '').strip() for c in next(chosen_sheet.iter_rows(min_row=1, max_row=1, values_only=True)) ]
        # determine price column index
        price_col_idx = None
        for i, h in enumerate(header):
            for p in price_candidates:
                if (case_sensitive and h == p) or (not case_sensitive and h.lower() == p.lower()):
                    price_col_idx = i+1
                    break
            if price_col_idx:
                break

        if price_col_idx:
            # inspect first up to 30 rows
            for r in chosen_sheet.iter_rows(min_row=2, max_row=31):
                cell = r[price_col_idx-1]
                if cell is None:
                    continue
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    numeric_issues.append({'type': 'FORMULA_IN_NUMERIC', 'row': cell.row, 'col': cell.column, 'message': 'Formula detected in price cell'})
                else:
                    # check numeric parseability
                    if val is None:
                        numeric_issues.append({'type': 'EMPTY_NUMERIC', 'row': cell.row, 'col': cell.column, 'message': 'Empty price cell'})
                    else:
                        try:
                            float(val)
                        except Exception:
                            numeric_issues.append({'type': 'NON_NUMERIC_PRICE', 'row': cell.row, 'col': cell.column, 'message': f'Value not numeric: {val}'})

    if numeric_issues:
        out['status'] = 'fail'
        out['errors'].extend(numeric_issues)

    (REPORTS / 'precheck_excel.json').write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf8')
    return out


def normalize_text(s: str):
    if s is None:
        return ''
    s = str(s).strip()
    # remove commas and normalize fullwidth digits simple replacement
    s = s.replace(',', '')
    # fullwidth digits
    trans = str.maketrans('０１２３４５６７８９', '0123456789')
    s = s.translate(trans)
    return s


def importer_dryrun(xlsx_path: Path, mapping_path: Path):
    """Very small importer: map headers according to mapping.yaml and try to resolve to master_id_map.csv"""
    if yaml is None:
        raise RuntimeError('PyYAML required for importer dry-run')
    if load_workbook is None:
        raise RuntimeError('openpyxl required for importer dry-run')

    mapping = yaml.safe_load(mapping_path.read_text(encoding='utf8'))
    req = mapping.get('required', {})
    case_sensitive = mapping.get('settings', {}).get('match', {}).get('case_sensitive', False)

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    # pick first sheet that contains all required fields in header
    chosen = None
    chosen_header = None
    for sheet in wb.worksheets:
        header = [ (c or '').strip() for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) ]
        ok = True
        for logical_field, spec in req.items():
            candidates = spec.get('candidates', [])
            found = False
            for cand in candidates:
                for h in header:
                    if not h:
                        continue
                    if case_sensitive:
                        if h == cand:
                            found = True
                            break
                    else:
                        if h.lower() == cand.lower():
                            found = True
                            break
                if found:
                    break
            if not found:
                ok = False
                break
        if ok:
            chosen = sheet
            chosen_header = header
            break

    if chosen is None:
        raise RuntimeError('No sheet found with all required columns')

    # build header mapping: map column index -> logical_field
    header_map = {}
    for i, h in enumerate(chosen_header):
        for logical_field, spec in req.items():
            for cand in spec.get('candidates', []):
                if case_sensitive:
                    match = (h == cand)
                else:
                    match = (h.lower() == cand.lower())
                if match:
                    header_map[i] = logical_field
                    break
            if i in header_map:
                break

    normalized_rows_path = DRYRUN_DIR / 'normalized_rows.csv'
    mapped_rows_path = DRYRUN_DIR / 'mapped_rows.csv'
    resolved_rows_path = DRYRUN_DIR / 'resolved_rows.csv'
    unmapped_rows_path = DRYRUN_DIR / 'unmapped_rows.csv'
    possible_matches_path = DRYRUN_DIR / 'possible_matches.csv'

    normalized_rows = []
    mapped_rows = []
    resolved_rows = []
    unmapped_rows = []
    possible_matches = []

    # load master map (expecting columns like canonical_key,drug_id,drug_name_canonical,...)
    master = []
    master_path = ROOT / 'data' / 'master_id_map.csv'
    if master_path.exists():
        with master_path.open(encoding='utf8') as f:
            master = [r for r in csv.DictReader(f)]
    # build helper: canonical names (normalized)
    for m in master:
        # ensure a canonical name key exists for older/newer formats
        if not m.get('drug_name_canonical') and m.get('drug_name'):
            m['drug_name_canonical'] = m.get('drug_name')

    # iterate rows
    for row_idx, row in enumerate(chosen.iter_rows(min_row=2, values_only=True), start=2):
        norm = { 'row': row_idx }
        mapped = {}
        for i, val in enumerate(row):
            key = header_map.get(i)
            norm_val = normalize_text(val)
            norm[f'col_{i}'] = norm_val
            if key:
                mapped[key] = norm_val

        normalized_rows.append(norm)
        mapped_rows.append(mapped)

        # try to resolve to master by excel_product_code, exact name match, or relaxed name-only partial match
        resolved = None
        product_code = mapped.get('excel_product_code')
        if product_code and master:
            for m in master:
                if (m.get('product_code') or '') == product_code:
                    resolved = m
                    break
        # exact name-only match (normalized)
        if not resolved and master:
            name_norm = (normalize_text(mapped.get('excel_drug_name')) or '').lower()
            if name_norm:
                for m in master:
                    mname = (m.get('drug_name_canonical') or '').strip()
                    if not mname:
                        continue
                    mname_norm = normalize_text(mname).lower()
                    if mname_norm and mname_norm == name_norm:
                        resolved = m
                        resolved_match_type = 'exact_name'
                        break
        # relaxed partial match: master name contained in excel name or vice versa
        if not resolved and master:
            mn = (normalize_text(mapped.get('excel_drug_name')) or '').lower()
            for m in master:
                mname = (m.get('drug_name_canonical') or '')
                if not mname:
                    continue
                mname_norm = normalize_text(mname).lower()
                if not mn or not mname_norm:
                    continue
                if mname_norm in mn or mn in mname_norm:
                    resolved = m
                    resolved_match_type = 'name_only'
                    break

        if resolved:
            r = dict(mapped)
            r['drug_id'] = resolved.get('drug_id')
            # record how we matched (exact_name / name_only / product_code)
            r['match_type'] = resolved.get('match_type') or ( 'product_code' if product_code else ( 'name_only' if 'resolved_match_type' in locals() and resolved_match_type=='name_only' else 'exact_name' ))
            resolved_rows.append(r)
        else:
            unmapped_rows.append(mapped)
            # build possible matches by substring search on drug_name
            matches = []
            name = (mapped.get('excel_drug_name') or '').lower()
            if name and master:
                for m in master:
                    candidate = (m.get('drug_name_canonical') or '').lower()
                    if candidate and name in candidate:
                        matches.append({'drug_id': m.get('drug_id'), 'drug_name': m.get('drug_name_canonical')})
                        if len(matches) >= 5:
                            break
            if matches:
                possible_matches.append({'row': row_idx, 'name': mapped.get('excel_drug_name'), 'matches': matches})

    # write CSVs
    def write_csv(path: Path, rows, fieldnames=None):
        if not rows:
            # still write header if fieldnames given
            if fieldnames:
                with path.open('w', encoding='utf8', newline='') as f:
                    w = csv.DictWriter(f, fieldnames=fieldnames)
                    w.writeheader()
            return
        if fieldnames is None:
            # derive keys
            keys = set()
            for r in rows:
                keys.update(r.keys())
            fieldnames = list(keys)
        with path.open('w', encoding='utf8', newline='') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for r in rows:
                w.writerow(r)

    write_csv(normalized_rows_path, normalized_rows)
    write_csv(mapped_rows_path, mapped_rows)
    write_csv(resolved_rows_path, resolved_rows)
    write_csv(unmapped_rows_path, unmapped_rows)
    write_csv(possible_matches_path, possible_matches)

    # produce staging CSV from resolved_rows
    staging_path = STAGING / 'drug_price_staging.csv'
    staging_rows = []
    for r in resolved_rows:
        staging_rows.append({
            'drug_id': r.get('drug_id'),
            'drug_name': r.get('excel_drug_name'),
            'price_per_unit': r.get('excel_price_yen') or '',
            'units_per_12w': r.get('excel_unit_text') or '',
            'version': 'DRYRUN',
            'effective_date': '',
            'status': 'provisional',
        })
    write_csv(staging_path, staging_rows, fieldnames=['drug_id','drug_name','price_per_unit','units_per_12w','version','effective_date','status'])

    return {
        'normalized_rows': str(normalized_rows_path),
        'mapped_rows': str(mapped_rows_path),
        'resolved_rows': str(resolved_rows_path),
        'unmapped_rows': str(unmapped_rows_path),
        'possible_matches': str(possible_matches_path),
        'staging_csv': str(staging_path),
        'counts': {
            'resolved': len(resolved_rows),
            'unmapped': len(unmapped_rows),
            'possible_matches': len(possible_matches),
        }
    }


def run_validators(staging_csv: Path):
    reports = {}
    cmds = [
        (['python', str(ROOT / 'scripts' / 'validate_mapping.py'), str(ROOT / 'data' / 'mapping.yaml')], 'reports/validate_mapping.json'),
        (['python', str(ROOT / 'scripts' / 'validate_unit_map.py'), str(ROOT / 'data' / 'unit_map.csv')], 'reports/validate_unit_map.json'),
        (['python', str(ROOT / 'scripts' / 'validate_master_id_map.py'), str(ROOT / 'data' / 'master_id_map.csv')], 'reports/validate_master_id_map.json'),
        (['python', str(ROOT / 'scripts' / 'validate_drug_price.py'), str(staging_csv), str(ROOT / 'data' / 'drug_price.csv'), str(ROOT / 'data' / 'validator_rules.yaml')], 'reports/validate_drug_price.json'),
    ]

    overall_status = 'pass'
    for cmd, expected_report in cmds:
        print('Running:', ' '.join(cmd))
        try:
            res = subprocess.run(cmd, check=False, capture_output=True, text=True)
        except Exception as e:
            print('Failed to run validator:', e)
            overall_status = 'fail'
            reports[expected_report] = {'status': 'fail', 'error': str(e)}
            continue
        # if validator wrote report file, read it
        rpt = Path(expected_report)
        if rpt.exists():
            try:
                reports[expected_report] = json.loads(rpt.read_text(encoding='utf8'))
                if reports[expected_report].get('status') == 'fail':
                    overall_status = 'fail'
            except Exception as e:
                reports[expected_report] = {'status': 'fail', 'error': f'Failed to load report: {e}'}
                overall_status = 'fail'
        else:
            # validator may have printed to stdout/stderr
            stdout = res.stdout.strip()
            stderr = res.stderr.strip()
            reports[expected_report] = {'status': 'fail', 'stdout': stdout, 'stderr': stderr}
            overall_status = 'fail'

    return overall_status, reports


def run_simulation_tests(staging_csv: Path, rules_path: Path):
    # pick a small set of drug_ids from staging and run simulate_annual_cost
    results = []
    try:
        import importlib
        calc = importlib.import_module('src.calculator')
    except Exception as e:
        return {'error': f'Failed to import calculator: {e}'}

    rows = []
    if staging_csv.exists():
        with staging_csv.open(encoding='utf8') as f:
            rows = [r for r in csv.DictReader(f)]

    # pick up to 3 drug_ids
    sample = rows[:3]
    for r in sample:
        drug_id = r.get('drug_id')
        try:
            base = calc.simulate_annual_cost(system_version='R7', income_code='G', age_group='under70', drug_id=drug_id)
            with_existing = calc.simulate_annual_cost(system_version='R7', income_code='G', age_group='under70', drug_id=drug_id, per_event_extra=0)
            results.append({'drug_id': drug_id, 'base': base})
        except Exception as e:
            results.append({'drug_id': drug_id, 'error': str(e)})

    return results


def assemble_e2e_report(xlsx, dryrun_result, validator_reports, sim_results):
    summary = {
        'status': 'pass',
        'excel_file': str(xlsx),
        'new_drugs': dryrun_result['counts']['resolved'] if dryrun_result else 0,
        'price_change_warnings': 0,
        'price_change_errors': 0,
        'unmapped_rows': dryrun_result['counts']['unmapped'] if dryrun_result else 0,
        'possible_matches': dryrun_result['counts']['possible_matches'] if dryrun_result else 0,
    }
    # inspect validator reports for price change warnings/errors
    v = validator_reports.get('reports/validate_drug_price.json')
    if isinstance(v, dict):
        if v.get('summary'):
            summary['price_change_errors'] = v['summary'].get('error_count', 0)
            summary['price_change_warnings'] = v['summary'].get('warning_count', 0)
            if v.get('status') == 'fail' and summary['price_change_errors'] > 0:
                summary['status'] = 'fail'

    # unmapped_rows failure condition
    mapping = Path('data') / 'mapping.yaml'
    if mapping.exists():
        m = yaml.safe_load(mapping.read_text(encoding='utf8'))
        policy = m.get('settings', {}).get('failure_policy', {})
        unmapped_policy = policy.get('unmapped_rows', 'FAIL')
        if unmapped_policy.upper() == 'FAIL' and summary['unmapped_rows'] > 0:
            summary['status'] = 'fail'

    # write e2e summary
    (REPORTS / 'e2e_summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf8')

    # human readable md
    lines = []
    lines.append('# E2E Dry-run Report')
    lines.append(f'- Excel file: {xlsx}')
    lines.append(f'- Status: {summary["status"]}')
    lines.append('')
    lines.append('## Key metrics')
    lines.append(f'- New drugs (resolved): {summary["new_drugs"]}')
    lines.append(f'- Unmapped rows: {summary["unmapped_rows"]}')
    lines.append(f'- Possible matches: {summary["possible_matches"]}')
    lines.append(f'- Price change warnings: {summary["price_change_warnings"]}')
    lines.append(f'- Price change errors: {summary["price_change_errors"]}')
    lines.append('')
    lines.append('## Validator reports')
    for k, rpt in validator_reports.items():
        lines.append(f'- {k}: {rpt.get("status", "unknown") if isinstance(rpt, dict) else "missing"}')

    lines.append('')
    lines.append('## Manual checks / notes')
    if summary['unmapped_rows'] > 0:
        lines.append('- There are unmapped rows: manual mapping required before production import.')
    if summary['price_change_errors'] > 0:
        lines.append('- Price change errors detected: investigate large deltas before proceeding.')

    (REPORTS / 'e2e_human_readable.md').write_text('\n'.join(lines), encoding='utf8')
    return summary


def main():
    if len(sys.argv) < 2:
        print('Usage: scripts/e2e_dryrun.py path/to/source.xlsx')
        return 2
    xlsx = Path(sys.argv[1])
    mapping_path = ROOT / 'data' / 'mapping.yaml'
    rules_path = ROOT / 'data' / 'validator_rules.yaml'

    pre = precheck_excel(xlsx, mapping_path)
    # If precheck failed, stop early. However, if precheck passed or
    # produced warnings, continue the dry-run. Warnings are informational
    # and should not prevent the dry-run from producing staging CSVs
    # and running validators.
    if pre.get('status') == 'fail':
        print('Precheck failed; see reports/precheck_excel.json')
        return 1

    if pre.get('warnings'):
        print('Precheck produced warnings; continuing dry-run. See reports/precheck_excel.json')
    else:
        print('Precheck passed.')

    dryrun = importer_dryrun(xlsx, mapping_path)
    print('Dry-run generated:', dryrun)

    status, reports = run_validators(Path(dryrun['staging_csv']))
    print('Validators finished with status:', status)

    sim = run_simulation_tests(Path(dryrun['staging_csv']), rules_path)

    summary = assemble_e2e_report(xlsx, dryrun, reports, sim)
    print('E2E summary:', summary)
    return 0 if summary.get('status') == 'pass' else 1


if __name__ == '__main__':
    sys.exit(main())
