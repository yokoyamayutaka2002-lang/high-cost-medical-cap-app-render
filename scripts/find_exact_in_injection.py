#!/usr/bin/env python3
import sys
from pathlib import Path
from openpyxl import load_workbook

NAME_CANDIDATES = ["品名", "名称", "製品名"]
SPEC_CANDIDATES = ["規格", "規　格", "規格・容量", "規格（g）", "規格・用法"]
PRICE_CANDIDATES = ["薬価", "薬価（円）", "金額", "点数"]


def detect_header_indices(row_values):
    idx = {"name": None, "spec": None, "price": None}
    for i, v in enumerate(row_values):
        if v in NAME_CANDIDATES:
            idx["name"] = i
        if v in SPEC_CANDIDATES:
            idx["spec"] = i
        if v in PRICE_CANDIDATES:
            idx["price"] = i
    return idx


def main():
    if len(sys.argv) < 3:
        print("Usage: find_exact_in_injection.py <target> <excel_path>")
        sys.exit(2)

    target = sys.argv[1]
    excel_path = Path(sys.argv[2])

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    print("sheet,row,品名,規格,薬価")
    found_any = False

    for sheet in wb.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header_row = None
        header_idx = None
        for idx, row in enumerate(rows, start=1):
            header_idx_candidate = detect_header_indices(row)
            if header_idx_candidate["name"] is not None and header_idx_candidate["price"] is not None:
                header_row = idx
                header_idx = header_idx_candidate
                break
        if header_row is None:
            continue

        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if r_idx <= header_row:
                continue
            try:
                name_val = row[header_idx["name"]] if header_idx["name"] < len(row) else None
                spec_val = row[header_idx["spec"]] if header_idx.get("spec") is not None and header_idx["spec"] < len(row) else None
                price_val = row[header_idx["price"]] if header_idx["price"] < len(row) else None
            except Exception:
                continue
            if name_val is None:
                continue
            name_str = str(name_val).strip()
            if name_str == target:
                spec_str = "" if spec_val is None else str(spec_val).strip()
                price_str = "" if price_val is None else str(price_val).strip()
                print(f"{sheet.title},{r_idx},{name_str},{spec_str},{price_str}")
                found_any = True

    if not found_any:
        print(f"No matches for: {target}")


if __name__ == '__main__':
    main()
