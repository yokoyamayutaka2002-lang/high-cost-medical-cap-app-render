#!/usr/bin/env python3
"""Extract biologic drug prices from MHLW injection Excel by exact name matches.
Usage:
  python scripts/get_biologic_price_from_excel.py data/biologic_drug_master_exact.csv data/source_excel/mhlw_drug_price_injection_2025-04.xlsx
Output: reports/biologic_drug_price_2025-04.csv
"""
import csv
import sys
import re
from pathlib import Path
from openpyxl import load_workbook

NAME_CANDIDATES = ["品名", "名称", "製品名"]
PRICE_CANDIDATES = ["薬価", "薬価（円）", "金額", "点数"]


def parse_master(path):
    items = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("exact_item_name") or "").strip()
            if name:
                items.append(name)
    return items


def detect_header(row_values):
    name_idx = None
    price_idx = None
    for i, v in enumerate(row_values):
        if v in NAME_CANDIDATES:
            name_idx = i
        if v in PRICE_CANDIDATES:
            price_idx = i
    return name_idx, price_idx


def normalize_price(val):
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    s = s.replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-"):
        return None
    try:
        f = float(s)
        return int(round(f))
    except Exception:
        return None


def main():
    if len(sys.argv) < 3:
        print("Usage: get_biologic_price_from_excel.py <master_csv> <source_excel>")
        sys.exit(2)

    master_csv = Path(sys.argv[1])
    excel_path = Path(sys.argv[2])
    out_path = Path("reports/biologic_drug_price_2025-04.csv")

    master_items = parse_master(master_csv)
    if not master_items:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["exact_item_name", "price_yen", "source_excel", "sheet", "row", "class", "route"])
        print("No items in master; wrote empty output.")
        sys.exit(0)

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    results = {}
    header_found_in_some_sheet = False

    for sheet in wb.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header_row_index = None
        header_map = {}
        for idx, row in enumerate(rows, start=1):
            name_idx, price_idx = detect_header(row)
            if name_idx is not None and price_idx is not None:
                header_row_index = idx
                header_map = {"name": name_idx, "price": price_idx}
                header_found_in_some_sheet = True
                break

        if header_row_index is None:
            continue

        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if r_idx <= header_row_index:
                continue
            name_val = None
            price_val = None
            try:
                name_val = row[header_map["name"]] if header_map["name"] < len(row) else None
                price_val = row[header_map["price"]] if header_map["price"] < len(row) else None
            except Exception:
                continue
            if name_val is None:
                continue
            name_str = str(name_val).strip()
            if name_str in results:
                continue
            if name_str in master_items:
                price_yen = normalize_price(price_val)
                results[name_str] = {
                    "exact_item_name": name_str,
                    "price_yen": price_yen if price_yen is not None else "",
                    "source_excel": str(excel_path),
                    "sheet": sheet.title,
                    "row": r_idx,
                    "class": "biologic",
                    "route": "injection",
                }

    if not header_found_in_some_sheet:
        print("ERROR: failed to detect required columns (品名/薬価) in any sheet.")
        sys.exit(3)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["exact_item_name", "price_yen", "source_excel", "sheet", "row", "class", "route"])
        for name in master_items:
            rec = results.get(name)
            if rec:
                w.writerow([
                    rec["exact_item_name"],
                    rec["price_yen"],
                    rec["source_excel"],
                    rec["sheet"],
                    rec["row"],
                    rec["class"],
                    rec["route"],
                ])

    found_set = set(results.keys())
    for name in master_items:
        if name not in found_set:
            print(f"WARN: not found in Excel: {name}")

    print(f"Wrote {out_path}")


if __name__ == '__main__':
    main()
