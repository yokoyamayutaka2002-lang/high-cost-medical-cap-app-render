#!/usr/bin/env python3
"""Lookup exact price_per_unit from 厚労省 Excel by drug_name and strength.

Rules:
- Exact-match on normalized text (strip, fullwidth->halfwidth letters/digits, lowercasing for ASCII letters).
- If exactly one matching row is found → return numeric price (float).
- If zero matches → raise LookupError
- If multiple matches → raise LookupError with explanation

This script also runs checks for the requested examples and prints results.
"""
from pathlib import Path
import csv
from openpyxl import load_workbook
import sys
import re


def normalize_cell(s: str) -> str:
    if s is None:
        return ''
    s = str(s).strip()
    # translate fullwidth digits and ascii-fullwidth letters to ascii
    trans_digits = str.maketrans('０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐqrstuvwxyzＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰQRSTUVWXYZ',
                                 '0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ')
    try:
        s = s.translate(trans_digits)
    except Exception:
        pass
    # normalize micro sign to mu and remove full-width spaces
    s = s.replace('µ', 'u').replace('μ', 'u').replace('\u3000', ' ')
    return s


def find_header_indices(header_row):
    # header_row: list of strings
    names = [ (h or '').strip() for h in header_row ]
    idx = {'name': None, 'strength': None, 'price': None}
    name_keys = ['品名', '医薬品名', '薬品名', '商品名', '製品名']
    strength_keys = ['規格', '含量', '規格・含量']
    price_keys = ['薬価', '価格', '単価']
    for i, h in enumerate(names):
        if not h:
            continue
        for k in name_keys:
            if k in h:
                idx['name'] = i
        for k in strength_keys:
            if k in h:
                idx['strength'] = i
        for k in price_keys:
            if k in h:
                idx['price'] = i
    return idx


def get_prices_for_name(target_drug_name: str, xlsx_path: Path):
    """Search sheets for rows where the 品名 contains target_drug_name (substring).

    For each matching row, extract mg from the 品名 string using regex
    (supports 'mg' and 'ｍｇ'). Group by extracted mg and for each mg value
    keep the first-appearance row's price. Rows without mg extraction are
    printed as WARN. If no rows match the target_drug_name at all, raise
    LookupError.

    Returns a list of dicts: {drug_name, strength_mg, price_yen, source_row_index}.
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    target_norm = normalize_cell(target_drug_name)

    # patterns
    mg_re = re.compile(r"(\d+)\s*[mｍ]g", flags=re.IGNORECASE)
    uju_re = re.compile(r"(?:u|Ｕ)\s*錠", flags=re.IGNORECASE)

    # per-target allowed mg whitelist (integers)
    mg_whitelist = {
        'テオフィリン徐放U錠': {100, 200, 400},
        'モンテルカスト錠': {5, 10},
        'プランルカスト錠': {225},
    }

    grouped = {}  # mg (int) -> {drug_name, strength_mg, price_yen, source_row_index}
    any_hit = False
    warn_rows = []

    for sheet in wb.worksheets:
        try:
            header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            continue
        idx = find_header_indices(header)
        if idx['name'] is None or idx['price'] is None:
            continue

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            cell_name_raw = row[idx['name']]
            cell_price_raw = row[idx['price']]
            cell_name = normalize_cell(cell_name_raw)
            if not cell_name:
                continue

            # --- selection logic per drug ---
            use_row = False
            # テオフィリン徐放U錠: require 品名 contains テオフィリン and contains U錠/Ｕ錠
            if target_norm == 'テオフィリン徐放U錠':
                if 'テオフィリン' in cell_name and uju_re.search(cell_name):
                    use_row = True

            # モンテルカスト錠: anchor 'モンテルカスト' (substring match)
            elif 'モンテルカスト' in target_norm:
                if 'モンテルカスト' in cell_name:
                    use_row = True

            # プランルカスト錠: anchor 'プランルカスト'
            elif 'プランルカスト' in target_norm:
                if 'プランルカスト' in cell_name:
                    use_row = True

            # fallback: substring match of target name in 品名
            else:
                if target_norm in cell_name:
                    use_row = True

            if not use_row:
                continue
            any_hit = True

            # try extract mg from 品名 first, then 規格 column if available
            m = mg_re.search(cell_name)
            strength_mg = None
            if m:
                try:
                    strength_mg = int(m.group(1))
                except Exception:
                    strength_mg = None

            # if not found in 品名, try 規格 column
            if strength_mg is None and idx.get('strength') is not None:
                strength_raw = row[idx['strength']]
                strength_text = normalize_cell(strength_raw)
                m2 = mg_re.search(strength_text)
                if m2:
                    try:
                        strength_mg = int(m2.group(1))
                    except Exception:
                        strength_mg = None

            if strength_mg is None:
                # cannot determine mg — warn and skip
                warn_rows.append({'sheet': sheet.title, 'row': row_idx, 'excel_item_name': cell_name})
                continue

            # filter by whitelist if present for this target
            allowed = mg_whitelist.get(target_norm)
            if allowed is not None and strength_mg not in allowed:
                print(f"INFO: extracted mg={strength_mg} not in whitelist for {target_drug_name}, skipping (sheet={sheet.title}, row={row_idx}, name={cell_name})")
                continue

            # parse price to float if possible
            try:
                price = float(str(cell_price_raw).replace(',', ''))
            except Exception:
                price = None

            # if mg not yet recorded, store this row (first-appearance rule)
            if strength_mg not in grouped:
                grouped[strength_mg] = {
                    'drug_name': target_drug_name,
                    'strength_mg': strength_mg,
                    'price_yen': price,
                    'source_row_index': row_idx,
                    'sheet': sheet.title,
                    'excel_item_name': cell_name,
                }

    # emit warnings
    for w in warn_rows:
        print('WARN: could not extract mg/規格 for', w)

    if not any_hit:
        raise LookupError(f'No rows matched target drug name: {target_drug_name}')

    # return as list sorted by numeric mg
    out = []
    for mg in sorted(grouped.keys()):
        g = grouped[mg]
        out.append({
            'drug_name': g['drug_name'],
            'strength_mg': g['strength_mg'],
            'price_yen': g['price_yen'],
            'source_row_index': g['source_row_index'],
            'row': g['source_row_index'],
            'sheet': g['sheet'],
            'excel_item_name': g['excel_item_name'],
        })

    return out


def main():
    xlsx = Path('data') / 'source_excel' / 'mhlw_drug_price_oral_2025-04.xlsx'
    targets = [
        'テオフィリン徐放U錠',
        'モンテルカスト錠',
        'プランルカスト錠',
    ]
    all_results = {}
    for t in targets:
        try:
            res = get_prices_for_name(t, xlsx)
            all_results[t] = res
        except LookupError as e:
            print(f'INFO: {e}')
            all_results[t] = []

    # print results in requested format
    for t, rows in all_results.items():
        print('\nResults for:', t)
        if not rows:
            print('  (no matches with mg extracted)')
            continue
        print('drug_name,strength_mg,excel_item_name,price_yen,sheet,row')
        for r in rows:
            print(f"{r['drug_name']},{r['strength_mg']},{r['excel_item_name']},{r['price_yen']},{r['sheet']},{r['row']}")

    # Write whitelist-passed results to CSV
    out_dir = Path('reports')
    out_dir.mkdir(parents=True, exist_ok=True)
    out_csv = out_dir / 'oral_drug_price_2025-04.csv'
    with out_csv.open('w', encoding='utf-8-sig', newline='') as fh:
        writer = csv.writer(fh)
        writer.writerow(['drug_name', 'strength_mg', 'price_yen', 'source_excel', 'sheet', 'row'])
        for t, rows in all_results.items():
            for r in rows:
                # only include rows that passed whitelist (function already filtered)
                writer.writerow([
                    r.get('drug_name'),
                    r.get('strength_mg'),
                    r.get('price_yen'),
                    xlsx.name,
                    r.get('sheet'),
                    r.get('row'),
                ])
    print(f"WROTE: {out_csv}")


if __name__ == '__main__':
    main()
