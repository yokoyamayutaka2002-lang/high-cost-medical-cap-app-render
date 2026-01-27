#!/usr/bin/env python3
"""Inspect the mhlw oral price Excel for rows whose 品名 contains 'U錠'.

Outputs:
- sheet name, row number, repr(品名), repr(規格), 薬価
- then groups by extracted (\d+)mg from 品名 and lists the first-appearance price per mg

Run from repository root with the same venv used elsewhere.
"""
from pathlib import Path
import re
from openpyxl import load_workbook


def normalize_cell(s):
    if s is None:
        return ''
    s = str(s).strip()
    # basic fullwidth → halfwidth digits/letters
    trans_digits = str.maketrans('０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐqrstuvwxyzＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰQRSTUVWXYZ',
                                 '0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ')
    try:
        s = s.translate(trans_digits)
    except Exception:
        pass
    s = s.replace('µ', 'u').replace('μ', 'u').replace('\u3000', ' ')
    return s


def find_header_indices(header_row):
    names = [ (h or '').strip() for h in header_row ]
    idx = {'name': None, 'strength': None, 'price': None}
    name_keys = ['品名', '医薬品名', '薬品名', '商品名', '製品名']
    strength_keys = ['規格', '含量', '規格・含量']
    price_keys = ['薬価', '価格', '単価']
    for i, h in enumerate(names):
        if not h:
            continue
        for k in name_keys:
            if k in h:
                idx['name'] = i
        for k in strength_keys:
            if k in h:
                idx['strength'] = i
        for k in price_keys:
            if k in h:
                idx['price'] = i
    return idx


def inspect_xlsx(xlsx_path: Path):
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    token = normalize_cell('U錠')
    # allow optional space and match either ASCII 'U' or fullwidth 'Ｕ'
    token_re = re.compile(r"(?:u|Ｕ)\s*錠", flags=re.IGNORECASE)
    mg_re = re.compile(r"(\d+)\s*mg", flags=re.IGNORECASE)

    matched_rows = []  # list of dicts

    # Print sheet names and header rows to help debug why 'U錠' may not be matching
    print('Sheets and header rows:')
    for sheet in wb.worksheets:
        try:
            header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            header = []
        print(f"- Sheet: {sheet.title} -> header: {header}")
    print('\nScanning rows that contain token:', repr(token))
    # Also print any rows that mention テオフィリン to help locate representations
    teo_token = normalize_cell('テオフィリン')
    teo_matches = []

    for sheet in wb.worksheets:
        try:
            header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            continue
        idx = find_header_indices(header)
        if idx['name'] is None or idx['price'] is None:
            continue

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name_raw = row[idx['name']]
            if name_raw is None:
                continue
            name = normalize_cell(name_raw)
            if not token_re.search(name):
                continue

            strength_raw = row[idx['strength']] if idx['strength'] is not None else None
            strength = normalize_cell(strength_raw) if strength_raw is not None else ''
            price_raw = row[idx['price']]
            # try to parse price to float
            try:
                price = float(str(price_raw).replace(',', ''))
            except Exception:
                price = price_raw

            matched_rows.append({
                'sheet': sheet.title,
                'row': row_idx,
                'excel_item_name': name,
                'strength_text': strength,
                'price': price,
            })

        # collect テオフィリン mention rows as well for debugging
        if teo_token:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                name_raw = row[idx['name']]
                if name_raw is None:
                    continue
                name = normalize_cell(name_raw)
                if teo_token in name:
                    teo_matches.append({'sheet': sheet.title, 'row': row_idx, 'name': name, 'spec': (row[idx['strength']] if idx['strength'] is not None else None), 'price': row[idx['price']]} )

    # print matched rows
    print('Matched rows containing "U錠":')
    if not matched_rows:
        print('  (none)')
    else:
        for r in matched_rows:
            print(f"Sheet={r['sheet']}, Row={r['row']}, name={repr(r['excel_item_name'])}, spec={repr(r['strength_text'])}, price={r['price']}")

    # print テオフィリン matches for inspection
    print('\nRows that mention "テオフィリン" in 品名 (for debugging):')
    if not teo_matches:
        print('  (none)')
    else:
        for r in teo_matches:
            spec = r['spec']
            try:
                spec_repr = repr(normalize_cell(spec))
            except Exception:
                spec_repr = repr(spec)
            print(f"Sheet={r['sheet']}, Row={r['row']}, name={repr(r['name'])}, spec={spec_repr}, price={r['price']}")

    # group by mg extracted from 品名
    grouped = {}
    warn_no_mg = []
    for r in matched_rows:
        m = mg_re.search(r['excel_item_name'])
        if not m:
            warn_no_mg.append(r)
            continue
        mg = int(m.group(1))
        if mg not in grouped:
            grouped[mg] = r

    if warn_no_mg:
        print('\nWARN: the following rows did not contain a (\\d+)mg match in 品名:')
        for r in warn_no_mg:
            print(f"  Sheet={r['sheet']}, Row={r['row']}, name={repr(r['excel_item_name'])}")

    # print grouped first-appearance prices
    print('\nFirst-appearance price per mg group:')
    if not grouped:
        print('  (no mg groups found)')
    else:
        for mg in sorted(grouped.keys()):
            r = grouped[mg]
            print(f"mg={mg}: price={r['price']} (sheet={r['sheet']}, row={r['row']}, name={repr(r['excel_item_name'])})")


def main():
    xlsx = Path('data') / 'source_excel' / 'mhlw_drug_price_oral_2025-04.xlsx'
    if not xlsx.exists():
        print('ERROR: expected Excel at', xlsx)
        return
    inspect_xlsx(xlsx)


if __name__ == '__main__':
    main()
