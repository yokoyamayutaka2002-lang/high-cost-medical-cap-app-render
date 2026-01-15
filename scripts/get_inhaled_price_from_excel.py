#!/usr/bin/env python3
"""Extract inhaled drug prices from 厚労省 topical Excel using exact-match master.

Outputs CSV: reports/inhaled_drug_price_2025-04.csv
Columns: exact_item_name,price_yen,source_excel,sheet,row,class,daily_inhalations
"""
from pathlib import Path
import csv
from openpyxl import load_workbook


def find_header_indices(header_row):
    names = [ (h or '') for h in header_row ]
    idx = {'name': None, 'strength': None, 'price': None}
    name_keys = ['品名', '医薬品名', '薬品名', '商品名', '製品名']
    strength_keys = ['規格', '含量', '規格含量']
    price_keys = ['薬価', '価格', '単価']
    for i, h in enumerate(names):
        if not h:
            continue
        for k in name_keys:
            if k in h:
                idx['name'] = i
        for k in strength_keys:
            if k in h:
                idx['strength'] = i
        for k in price_keys:
            if k in h:
                idx['price'] = i
    return idx


def load_master(path: Path):
    rows = []
    with path.open('r', encoding='utf-8') as fh:
        reader = csv.DictReader(fh)
        for r in reader:
            rows.append(r)
    return rows


def get_inhaled_prices_exact(master_csv: Path, xlsx: Path, out_csv: Path):
    wb = load_workbook(filename=str(xlsx), data_only=True)
    masters = load_master(master_csv)

    results = []
    for m in masters:
        exact = m.get('exact_item_name')
        daily = m.get('daily_inhalations')
        cls = m.get('class', '')
        if exact is None:
            print('WARN: master missing exact_item_name', m)
            continue
        found = False
        for sheet in wb.worksheets:
            try:
                header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
            except StopIteration:
                continue
            idx = find_header_indices(header)
            if idx['name'] is None or idx['price'] is None:
                continue

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                cell = row[idx['name']]
                if cell is None:
                    continue
                # strict exact equality after stripping surrounding whitespace
                try:
                    cell_text = str(cell)
                except Exception:
                    cell_text = cell
                if cell_text.strip() != exact:
                    continue

                # parse price
                price_raw = row[idx['price']]
                try:
                    price = float(str(price_raw).replace(',', ''))
                except Exception:
                    price = None

                results.append({
                    'exact_item_name': exact,
                    'price_yen': price,
                    'source_excel': xlsx.name,
                    'sheet': sheet.title,
                    'row': row_idx,
                    'class': cls,
                    'daily_inhalations': int(daily) if daily else None,
                })
                found = True
                break
            if found:
                break
        if not found:
            print('WARN: price not found for exact_item_name=', exact)

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    # Write results but skip duplicate exact_item_name entries (keep first appearance)
    seen = set()
    with out_csv.open('w', encoding='utf-8-sig', newline='') as fh:
        writer = csv.writer(fh)
        writer.writerow(['exact_item_name', 'price_yen', 'source_excel', 'sheet', 'row', 'class', 'daily_inhalations'])
        for r in results:
            exact = r.get('exact_item_name')
            if exact in seen:
                continue
            writer.writerow([r.get('exact_item_name'), r.get('price_yen'), r.get('source_excel'), r.get('sheet'), r.get('row'), r.get('class'), r.get('daily_inhalations')])
            seen.add(exact)

    return results


def main():
    repo = Path('.')
    master_csv = repo / 'data' / 'inhaled_drug_master_exact.csv'
    xlsx = repo / 'data' / 'source_excel' / 'mhlw_drug_price_topical_2025-04.xlsx'
    if not xlsx.exists():
        src = repo / 'data' / 'source_excel'
        candidates = [p for p in src.iterdir() if p.is_file() and 'topical' in p.name.lower() and '2025-04' in p.name]
        if candidates:
            xlsx = candidates[0]
            print('INFO: using topical Excel file:', xlsx.name)
    out_csv = repo / 'reports' / 'inhaled_drug_price_2025-04.csv'
    res = get_inhaled_prices_exact(master_csv, xlsx, out_csv)
    print('WROTE:', out_csv)
    for r in res:
        print(r)


if __name__ == '__main__':
    main()
